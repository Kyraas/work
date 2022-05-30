# -*- coding: utf-8 -*-
from sqlite3 import connect, Error
from openpyxl import load_workbook
from re import match
from os import path
import sys

# Подключаемся к БД
def open_database(file_path=None, login=None, new_login=None,
                    password=None, new_password=None, test_info=None,
                    name=None, time=None, tries=None, get_pass=None,
                    criterion=None, mark=None):
    try:
        # database_path = resource_path("data_files/Database.db")
        # con = connect(database_path)
        con = connect("Database.db")
        cur = con.cursor()
        if new_login is not None:    # смена логина
            return change_login(con, cur, new_login)
        elif new_password is not None:  # смена пароля
            return change_password(con, cur, password, new_password)
        elif login is not None: # проверка логина
            return check_login(cur, login)
        elif password is not None:
            return change_password(con, cur, password)
        elif file_path is not None: # обновление теста
            delete_data(con, cur)
            return open_file(con, cur, file_path)
        elif test_info is not None:
            return get_config(cur)
        elif name is not None or time is not None or tries is not None:
            return set_config(con, cur, name=name,
                                time=time, tries=tries)
        elif get_pass is not None:
            return get_password(cur)
        elif criterion is not None:
            return get_criterion(cur)
        elif mark is not None:
            return set_criterion(con, cur, mark=mark)
        else:   # получение данных
            return get_questions(cur), get_answers(cur)
    except Error as error:
        return "Ошибка: " + str(error)
    finally:
        if con:
            cur.close()
            con.close()

# Открываем Excel-файл
def open_file(con, cur, file_path):
    try:
        test_file = load_workbook(file_path)
        parse_excel(test_file, con, cur)
        return "Загрузка завершена."
    except FileNotFoundError:
        return f"Ошибка: Файл по указанному пути {file_path} не найден."
    finally:
        test_file.close()


# Получаем данные из файла
def parse_excel(test_file, con, cur):
    answers = []
    correct_answer = 0
    qnum = 1
    n = 0
    worksheet = test_file.active
    end = worksheet.max_row

    for cell in worksheet['A']:
        line = str(cell.value)
        result = match(r'Вопрос', line)
        if result:
            if answers:
                correct_answer = answers.pop()
                qnum += 1

            question = (qnum, line)
            n = set_answers(n, qnum-1, answers, correct_answer, con, cur)
            insert_question(con, cur, question)
            answers = []
        elif cell.row == end:   # если последняя строка
            correct_answer = line   # захватываем последнюю строку файла
            set_answers(n, qnum, answers, correct_answer, con, cur)
        else:
            answers.append(line)


# Получаем ответы к вопросам
def set_answers(n, question_id, answers, correct, con, cur):
    for i in range(len(answers)):
        num = "0" + str(n+i+1)
        if str(i+1) in correct:
            answer = (num, i+1, answers[i], question_id, 1)
            insert_answer(con, cur, answer)
        else:
            answer = (num, i+1, answers[i], question_id, 0)
            insert_answer(con, cur, answer)
    n += len(answers)
    return n


# Удаляем старые данные из БД
def delete_data(con, cur):
    cur.execute("DELETE FROM questions")
    cur.execute("DELETE FROM answers")
    con.commit()


# Вставляем вопрос в БД
def insert_question(con, cur, data):
    cur.execute("INSERT INTO questions VALUES (?, ?)", data)
    con.commit()


# Вставляем ответ в БД
def insert_answer(con, cur, data):
    cur.execute("INSERT INTO answers VALUES (?, ?, ?, ?, ?)", data)
    con.commit()


# Получаем вопросы из БД
def get_questions(cur):
    cur.execute("SELECT question FROM questions")
    result = cur.fetchall()
    return [", ".join(row) for row in result]


# Получаем ответы из БД
def get_answers(cur):
    cur.execute("SELECT question_id, answer, correct, number FROM answers")
    result = cur.fetchall()
    num = set()
    for row in result:
        num.add(row[0])
    n = len(num)
    k=0

    answers_options = [["".join(answer[1]) for answer in result
                        if int(answer[0]) == i] for i in range(1, n+1)]
    # correct_answers = [["".join(answer[1]) for answer in result
    #                     if int(answer[0]) == i and int(answer[2]) == 1]
    #                     for i in range(1, n+1)]
    correct_answers = [[int("".join(answer[3])) for answer in result
                        if int(answer[0]) == i and int(answer[2]) == 1]
                        for i in range(1, n+1)]

    for answer in result:
        if int(answer[2]) == 1:
            k+=1
    
    return answers_options, correct_answers, k


# Получаем логин из БД
def check_login(cur, login):
    result = cur.execute("SELECT login FROM admin").fetchall()
    cur_log = "".join(result[0])
    if login != cur_log:
        return "Неверный логин!"
    else:
        return ""


# Меняем логин в БД
def change_login(con, cur, login):
    cur.execute("UPDATE admin SET login = ? WHERE id=0", (str(login),))
    con.commit()
    return "Логин сохранён."


# Получаем и меняем пароль в БД
def change_password(con, cur, old, new=None):
    result = cur.execute("SELECT password FROM admin").fetchall()
    cur_passwrd = "".join(result[0])
    if old == cur_passwrd:
        if new:
            cur.execute("UPDATE admin SET password = ? WHERE id=0", (str(new),))
            con.commit()
            return "Новый пароль сохранён."
        else:
            return ""
    else:
        return "Неверный пароль!"


# Получение пароля из БД
def get_password(cur):
    result = cur.execute("SELECT password FROM admin").fetchall()
    cur_passwrd = "".join(result[0])
    return cur_passwrd


# Изменение информации о тесте
def set_config(con, cur, name=None, time=None, tries=None):
    if name is not None:
        cur.execute("UPDATE configuration SET student_name = ? WHERE id=0", (str(name),))
    if time is not None:
        cur.execute("UPDATE configuration SET time_limit = ? WHERE id=0", (int(time),))
    if tries is not None:
        cur.execute("UPDATE configuration SET tries = ? WHERE id=0", (int(tries),))
    con.commit()
    return "Данные успешно изменены."


# Получение информации о тесте
def get_config(cur):
    result = cur.execute("SELECT student_name, time_limit, tries FROM configuration").fetchall()
    return result[0]


# Получение критериев оценки теста
def get_criterion(cur):
    result = cur.execute("SELECT num, mark, points FROM evaluation_criterion").fetchall()
    return result


# Изменение критериев оценки теста
def set_criterion(con, cur, mark=None):
    if mark is not None:
        mark5, mark4, mark3 = mark
        if mark5 is not None:
            cur.execute("UPDATE evaluation_criterion SET points = ? WHERE num = 5", (mark5,))
        if mark4 is not None:
            cur.execute("UPDATE evaluation_criterion SET points = ? WHERE num = 4", (mark4,))
        if mark3 is not None:
            cur.execute("UPDATE evaluation_criterion SET points = ? WHERE num = 3", (mark3,))
        con.commit()
        return "Данные успешно изменены."


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = path.abspath(".")

    return path.join(base_path, relative_path)