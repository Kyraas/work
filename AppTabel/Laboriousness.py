# Хранение данных о сотрудниках в файле, возможность редактирования этого файла
import openpyxl
import re
from excelfiles import create_new, save_file
from vacation import get_days

month_dict = {
    "01": "Январь",
    "02": "Февраль",
    "03": "Март",
    "04": "Апрель",
    "05": "Май",
    "06": "Июнь",
    "07": "Июль",
    "08": "Август",
    "09": "Сентябрь",
    "10": "Октябрь",
    "11": "Ноябрь",
    "12": "Декабрь"
}

columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']

def read_file(name):
    try:
        with open("employees.txt", "r", encoding='utf-8') as f:
            for line in f:
                if name in line:
                    line = line.strip().split()
                    line = line[3:]
                    stav = line[-1:]
                    line = line[:-1]
                    dol = " ".join(line)
                    return dol, *stav
    except FileNotFoundError:
        print("Файл 'employees.txt' не найден.")
        return False

def get_date(string):
    date_order = string.group()
    day = int(date_order[:2])
    month_num = date_order[3:5]
    year = date_order[6:]
    month_str = month_dict.get(month_num)
    return day, month_str, year

def get_hours(string):
    hours = string.replace(',', '.') # замена запятой на точку
    hours = re.search(r'\d*\.*\d*', hours)
    hours = hours.group()
    return float(hours) # находим только действительные числа

def get_month_hours(sheet):
    for cell in sheet['2']:
        if cell.value == "Часов в месяце":
            k = cell.column
            kcol = openpyxl.utils.get_column_letter(k)
            month_hours = float(sheet[f'{kcol}3'].value)
    return month_hours

def get_cur_month(s):
    month_list = list(month_dict.values())  # преобразуем словарь в список месяцев, для поиска текущего месяца
    for month in month_list:
        if s == month.lower():
            return month

def get_sum(sheet, row):
    s = 0
    for i in columns:
        k = sheet[f'{i}{row}'].value
        if k != None:
            s += k
    return s

def open_file(name, month):
    wb = False
    process = True
    employee = True
    try:
        wb = openpyxl.load_workbook(f"Табель 2022 {name}.xlsx", data_only=True) # открываем в режиме data_only для получения результата вычисления формул
        sheet = wb[f'{month}']
        for cell in sheet['2']:
            if cell.value == "Всего":
                col = cell.column + 1
                col = openpyxl.utils.get_column_letter(col)
        max_rows = sheet.max_row
        fact_hours = sheet[f'{col}{max_rows-3}'].value
        month_hours = get_month_hours(sheet)

        if fact_hours == month_hours:

            print(f"Табель сотрудника {name} на {month} уже заполнен.")
            # ui.update_idletasks()

            wb.close()
            wb = False
        else:
            wb.close()
            wb = openpyxl.load_workbook(f"Табель 2022 {name}.xlsx")
    except FileNotFoundError:
        print(f"Файл 'Табель 2022 {name}.xlsx' не найден. Создаем новый...")
        process = create_new(f"Табель 2022 {name}.xlsx")
        if process == False:
            quit()
        print("Вводим данные в новый файл...")
        wb = openpyxl.load_workbook(f"Табель 2022 {name}.xlsx")
        sheet = wb['Январь']    # первый лист
        sheet['A1'] = name

        employee = read_file(name)
        if employee == False:
            input("Нажмите Enter, чтобы завершить программу.")
            exit()
        elif employee != None:
            for cell in sheet['3']:
                if cell.value == "Должность":
                    cell.value = employee[0]
                elif cell.value == 1:
                    cell.value = int(employee[1])
                    print("Сохраняем внесенные изменения...")
                    wb.save(f"Табель 2022 {name}.xlsx")
                    break
        else:
            print("ФИО сотрудника не найдено в списке.")
            wb.close()
            wb = False
    finally:
        if (process == False) or (employee == False):
            quit()
        return wb

def fill_cell(sheet, hours, sum_num, col, row, max_hours):
    if hours > max_hours:   # если оставшихся часов больше 8
        n = max_hours - sum_num
        hours = hours - n
        sheet[f'{col}{row}'] = float(n)
    elif hours + sum_num <= max_hours:  # если сумма меньше 8
        sheet[f'{col}{row}'] = float(hours)
        hours = 0
    else:
        n = max_hours - sum_num
        hours = hours - n
        sheet[f'{col}{row}'] = float(n)
    return hours

def insert_hours(sheet, name, month, start_col, hours, start_day=1):
    dict_days, err = get_days(name, month)
    day_row = [5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29,	31,	33,	35,	37,	39,	41,	43,	45,	47,	49,	51,	53,	55,	57,	59,	61,	63,	65]

    if err == True:
        input("Нажмите Enter, чтобы завершить программу.")
        quit()

    if type(start_col) == int:  # если получен номер стобца, то получаем букву
        column = openpyxl.utils.get_column_letter(start_col)
    else:
        column = start_col

    for day, row in enumerate(day_row, start=1):    # цикл по рабочим дням
        sum_num = get_sum(sheet, row)   # получаем сумму часов в дне
        if start_day > day: # если задана стартовая дата, то пропускаем дни, до этой даты
            pass
        elif (day in dict_days["work"]) and (sum_num != 8):  # если день рабочий и есть еще неотработанные в этот день часы
            hours = fill_cell(sheet, hours, sum_num, column, row, 8)
        elif (day in dict_days["pre_holiday"]) and (sum_num != 7):  # если день рабочий и есть еще неотработанные в этот день часы
            hours = fill_cell(sheet, hours, sum_num, column, row, 7)
        else:
            pass

def check_date(cell, sheet, name, cur_month, col, hours, check_hours):
    date_order = re.search(r'\d\d\.\d\d\.\d\d\d\d', str(cell.value))    # ищем дату в заказ-наряде (формат дд.мм.гггг)
    if date_order != None:  # если есть дата в заказ-наряде
        day, month, year = get_date(date_order) # преобразуем строку в числа
        if (year == "2022") and (month == cur_month):
            insert_hours(sheet, name, cur_month, col, hours, day)
        else:
            insert_hours(sheet, name, cur_month, col, hours, 1)
    else:   # если заказ-наряд не содержит дату
        insert_hours(sheet, name, cur_month, col, hours, 1)

    check_hours += hours
    return check_hours

def main():

    string = ""
    cur_month = ""
    name = ""
    hours = ""
    full_hours = 0.0
    check_hours = 0.0
    row_start = 0

    try:
        laborious = openpyxl.load_workbook("Трудоемкость за январь.xlsx")
        department = laborious['17 департамент']    # лист с нашим департаментом
        max_col = department.max_column
    except FileNotFoundError:
        print("Файл 'Трудоемкость за январь.xlsx' не найден.")
        quit()

    for cell in department['B']:
        string = str(cell.value)
        if (cur_month == "") and (string != "None"): # получим текущий месяц
            cur_month = get_cur_month(string)

        if string[:2] == '17':
            name = str(cell.value)  # получаем ФИО сотрудника
            name = re.search(r'\D+', name).group() # убираем числа из ФИО
            hours = department.cell(row=cell.row, column=max_col).value # получаем часы из соседней колонки 'С'
            check_hours = 0.0

            if (hours != None) and (hours != '0ч'):   # если у сотрудника есть суммарные рабочие часы
                timesheet = open_file(name, cur_month) # открываем\создаем именной табель
                if timesheet == False:
                    continue
                
                first = timesheet['Январь']
                sheet = timesheet[f'{cur_month}']
                max_rows = sheet.max_row
                full_hours = get_hours(hours) # преобразуем текст в действительное число
                row_start = int(cell.row)
                month_hours = get_month_hours(sheet)  # получение общих рабочих часов
                general = month_hours - full_hours
                print('\n',name, full_hours)

                
        elif (int(cell.row) != row_start) and (row_start != 0) and (timesheet != False):   # получаем заказ-наряды под каждого человека
            hours = department.cell(row=cell.row, column=max_col).value
            if (hours != None) and (hours != '0ч'):   # если есть рабочие часы на заказ-наряд
                hours = get_hours(hours)
                print(str(cell.value), hours)  # заказ-наряд

                for col in columns[1:]:
                    if first[f'{col}3'].value == None:   # если колонка пустая
                        first[f'{col}3'] = str(cell.value)   # вписываем заказ-наряд
                        sheet[f'{col}{max_rows-1}'] = hours  # вписываем часы заказ-наряда
                        check_hours = check_date(cell, sheet, name, cur_month, col, hours, check_hours)
                        break
                    elif first[f'{col}3'].value == str(cell.value):
                        sheet[f'{col}{max_rows-1}'] = hours
                        check_hours = check_date(cell, sheet, name, cur_month, col, hours, check_hours)
                        break
                    else:
                        continue

                if (check_hours == full_hours) and (full_hours != 0): # заполнение общих рабочих часов
                    print("Заполняем табель...")
                    insert_hours(sheet, name, cur_month, 5, general, 1)
                    print(f"Сохраняем и закрываем 'Табель 2022 {name}.xlsx' ...")
                    try:
                        timesheet.save(f"Табель 2022 {name}.xlsx")  # закрываем ранее открытый табель
                        timesheet.close()
                        save_file(f"Табель 2022 {name}.xlsx")
                    except PermissionError:
                        print(f"Ошибка доступа, пожалуйста закройте файл 'Табель 2022 {name}.xlsx'. ")
                        timesheet.close()

    laborious.close()
    print("Заполнение табелей завершено.")