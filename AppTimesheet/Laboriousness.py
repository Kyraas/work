import openpyxl
import re
from excelfiles import *
from pcalendar import *

employees = {
    "Бондаренко Мария Алексеевна": ["Инженер-стажер", "1"],
    "Михайлов Даниил Сергеевич": ["Инженер", "1"],
    "Перетягин Николай Александрович": ["Начальник отдела", "1"],
    "Поддубский Александр Анатольевич": ["Ведущий инженер", "1"],
    "Федорищев Иван Васильевич": ["Инженер", "1"],
    "Дрявичев Иван Олегович": ["Инженер-стажер", "0,5"],
    "Рыженко Сергей Викторович": ["Директор департамента", "1"],
    "Тараканов Иван Игоревич": ["Инженер", "1"]
}

month_dict = {
    "01": "Январь",
    "02": "Февраль",
    "03": "Март",
    "04": "Апрель",
    "05": "Май",
    "06": "Июнь",
    "07": "Июль",
    "08": "Август",
    "09": "Сентябрь",
    "10": "Октябрь",
    "11": "Ноябрь",
    "12": "Декабрь"
}

columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']

def get_date(string):
    date_order = string.group()
    day = int(date_order[:2])
    month_num = date_order[3:5]
    year = date_order[6:]
    month_str = month_dict.get(month_num)
    return day, month_str, year

def get_hours(string):
    hours = string.replace(',', '.') # замена запятой на точку
    hours = re.search(r'\d*\.*\d*', hours)
    hours = hours.group()
    return float(hours) # находим только действительные числа

def get_general_hours(sheet, hours):
    for cell in sheet['2']:
        if cell.value == "Часов в месяце":
            k = cell.column
            kcol = openpyxl.utils.get_column_letter(k)
            month_hours = float(sheet[f'{kcol}3'].value)
    general = month_hours - hours
    return general

def get_cur_month(s):
    month_list = list(month_dict.values())  # преобразуем словарь в список месяцев, для поиска текущего месяца
    for month in month_list:
        if s == month.lower():
            return month

def get_sum(sheet, row):
    s = 0
    for i in columns:
        k = sheet[f'{i}{row}'].value
        if k != None:
            s += k
    return s

def open_file(name):
    try:
        wb = openpyxl.load_workbook(f"Табель 2022 {name}.xlsx", keep_vba=True, data_only=True)
        print(f"Успешно открыт файл 'Табель 2022 {name}.xlsx'.")
    except FileNotFoundError:
        print(f"Файл 'Табель 2022 {name}.xlsx' не найден. Создаем новый...")
        create_new(f"Табель 2022 {name}.xlsx")
        print("Вводим данные в новый файл...")
        wb = openpyxl.load_workbook(f"Табель 2022 {name}.xlsx")
        sheet = wb['Январь']    # первый лист
        sheet['A1'] = name
        if name in employees:
            for cell in sheet['3']:
                if cell.value == "Должность":
                    cell.value = employees[name][0]
                elif cell.value == 1:
                    cell.value = int(employees[name][1])
            print("Сохраняем внесенные изменения...")
            wb.save(f"Табель 2022 {name}.xlsx")
            print("Пробуем снова ввести данные...")
            open_file(name)
        else:
            print("ФИО сотрудника не найдено в списке.")
    finally:
        return wb

def insert_hours(sheet, month, start_col, hours, day=1):
    dict_days = get_days(month)
    day_row = [5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29,	31,	33,	35,	37,	39,	41,	43,	45,	47,	49,	51,	53,	55,	57,	59,	61,	63,	65]
    if type(start_col) == int:
        column = openpyxl.utils.get_column_letter(start_col)
    else:
        column = start_col

    for cell in sheet['2']:
        if cell.value == "Всего":
            check_col = cell.column + 1
            check_col = openpyxl.utils.get_column_letter(check_col)

    for index, value in enumerate(day_row, start=1):
        sum_num = get_sum(sheet, value)
        if day > index: # если задана стартовая дата, то пропускаем дни, до этой даты
            pass
        elif (index in dict_days["work"]) & (sum_num != 8):  # если день рабочий и есть еще неотработанные в этот день часы
            if hours > 8:
                n = 8 - sum_num
                hours = hours - n
                sheet[f'{column}{value}'] = float(n)
            elif hours > 0:
                if hours + sum_num <= 8:
                    sheet[f'{column}{value}'] = float(hours)
                    hours = 0
                else:
                    n = 8 - sum_num
                    hours = hours - n
                    sheet[f'{column}{value}'] = float(n)
            else:
                break
        else:
            pass

string = ""
cur_month = ""
name = ""
hours = ""
full_hours = 0.0
check_hours = 0.0
row_start = 0
j = 0
order = ""
date_order = ""

laborious = openpyxl.load_workbook("Трудоемкость за январь.xlsx")
department = laborious['17 департамент']    # лист с нашим департаментом
max_col = department.max_column

for cell in department['B']:
    string = str(cell.value)
    if cur_month == "": # получим текущий месяц
        if string != "None":
            cur_month = get_cur_month(string)

    if string[:2] == '17':
        name = str(cell.value)  # получаем ФИО сотрудника
        name = re.search(r'\D+', name).group() # убираем числа из ФИО
        hours = department.cell(row=cell.row, column=max_col).value # получаем часы из соседней колонки 'С'
        check_hours = 0.0
        j = 1

        if (hours != None) & (hours != '0ч'):   # если у сотрудника есть суммарные рабочие часы
            timesheet = open_file(name) # открываем\создаем именной табель
            sheet = timesheet[f'{cur_month}']
            max_rows = sheet.max_row
            full_hours = get_hours(hours) # преобразуем текст в действительное число
            row_start = int(cell.row)
            general = get_general_hours(sheet, full_hours)  # получение общих рабочих часов
            print(name, full_hours)
            
    elif (int(cell.row) != row_start) & (row_start != 0):   # получаем заказ-наряды под каждого человека
        hours = department.cell(row=cell.row, column=max_col).value
        if (hours != None) & (hours != '0ч'):   # если есть рабочие часы на заказ-наряд
            hours = get_hours(hours)
            print(str(cell.value), hours)  # заказ-наряд
            sheet[f'{columns[j]}3'] = str(cell.value)   # вписываем заказ-наряд
            sheet[f'{columns[j]}{max_rows-1}'] = hours  # вписываем часы заказ-наряда

            date_order = re.search(r'\d\d\.\d\d\.\d\d\d\d', str(cell.value))    # ищем дату в заказ-наряде (формат дд.мм.гггг)
            if date_order != None:  # если есть дата в заказ-наряде
                day, month, year = get_date(date_order) # преобразуем строку в числа
                print(day, month, year)
                if year == "2022":
                    insert_hours(sheet, cur_month, columns[j], hours, day)
                else:
                    insert_hours(sheet, cur_month, columns[j], hours, 1)
            else:
                insert_hours(sheet, cur_month, columns[j], hours, 1)
            
            check_hours += hours
            j += 1  # следующая колонка в excel-файле

            if (check_hours == full_hours) & (full_hours != 0): # заполнение общих рабочих часов
                insert_hours(sheet, cur_month, 5, general, 1)
                print(f"Сохраняем и закрываем 'Табель 2022 {name}.xlsx' ...")
                timesheet.save(f"Табель 2022 {name}.xlsx")  # закрываем ранее открытый табель
                timesheet.close()

laborious.close()
