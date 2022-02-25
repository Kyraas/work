# Производственный календарь скачан с сайта: http://www.formy-i-blanki.ru/proizvodstvennyj_kalendar/2022
import openpyxl
import os

def set_year(year):
    fname = os.path.abspath(f"Производственный-календарь-{year}.xlsx")   # получаем абсолютный путь
    wb = openpyxl.load_workbook(fname)
    calendar = wb['Календарь']    # лист с нашим департаментом
    return calendar

def get_days(month, year=2022):
    calendar = set_year(year)   # получаем календарь нужного месяца
    rows = calendar.max_row
    cols = calendar.max_column

    for i in range(1, rows + 1):    # находим нужный месяц
        for j in range(1, cols + 1):
            if month == calendar.cell(row = i, column = j).value:
                row = i + 1
                col = j

    month_dict = {}
    work, pre_holiday = [], []  # рабочие и предпраздничные дни
    # hours = 0.0
    for j in range(col, col + 6):
        for i in range(row, row + 7):
            c = calendar.cell(row = i, column = j)
            if c.value != None: # если есть число
                if type(c.fill.start_color.index) == str:   # рабочие дни
                    work.append(c.value)
                elif c.fill.start_color.index == 0: # предпраздничные дни
                    pre_holiday.append(c.value)
    month_dict["work"] = work
    month_dict["pre_holiday"] = pre_holiday
    return month_dict
    